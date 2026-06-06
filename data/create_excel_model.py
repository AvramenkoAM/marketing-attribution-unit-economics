"""Creates unit_economics_model.xlsx with 3 sheets: Inputs, Unit Economics, Scenarios."""
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent.parent / "unit_economics_model.xlsx"

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
TEAL    = "FF2a9d8f"
NAVY    = "FF264653"
ORANGE  = "FFe76f51"
YELLOW  = "FFe9c46a"
LGRAY   = "FFf0f4f8"
WHITE   = "FFFFFFFF"
BLACK   = "FF1a202c"

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def val_cell(ws, row, col, v, fmt=None, bg=None):
    c = ws.cell(row=row, column=col, value=v)
    if fmt:  c.number_format = fmt
    if bg:   c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right")
    return c

def label(ws, row, col, v, bold=False):
    c = ws.cell(row=row, column=col, value=v)
    c.font = Font(bold=bold, size=10)
    return c

def thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CHANNELS  = ["Google Ads", "Facebook Ads", "SEO", "Email"]
CH_COLORS = [TEAL, ORANGE, NAVY, YELLOW]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: INPUTS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Inputs"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 32
for col in ["B","C","D","E"]: ws1.column_dimensions[col].width = 16
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 22

# Title
ws1.merge_cells("A1:E1")
hdr(ws1, 1, 1, "Unit Economics Model — Marketing Channel Inputs", NAVY, WHITE, size=13)

# Channel headers
hdr(ws1, 3, 1, "Parameter",         LGRAY, BLACK, bold=True)
for i, (ch, clr) in enumerate(zip(CHANNELS, CH_COLORS)):
    hdr(ws1, 3, i+2, ch, clr, WHITE)

PARAMS = [
    ("Monthly Budget ($)",        [8000,  6000,  2000,  500],   "#,##0",    "Monthly marketing spend allocated to channel"),
    ("New Customers / Month",     [137,   86,    162,   53],     "#,##0",    "Average monthly customer acquisitions"),
    ("CAC ($)",                   [78.8,  88.9,  14.9,  10.7],  "#,##0.00", "Cost per Acquired Customer = Spend / New Customers"),
    ("Avg Monthly ARPU ($)",      [48.0,  39.0,  55.0,  43.0],  "#,##0.00", "Average Revenue per User per Month"),
    ("Gross Margin (%)",          [0.72,  0.70,  0.76,  0.74],  "0.0%",     "Revenue remaining after COGS"),
    ("Monthly Churn Rate (%)",    [0.055, 0.082, 0.038, 0.028], "0.0%",     "% of customers lost per month"),
    ("LTV ($)",                   ["=C5*C6/C7", "=D5*D6/D7", "=E5*E6/E7", "=F5*F6/F7"],
                                                                "#,##0",    "LTV = ARPU × Margin / Monthly Churn"),
    ("LTV / CAC Ratio",           ["=C8/C4", "=D8/D4", "=E8/E4", "=F8/F4"],
                                                                "0.0x",     "Healthy: > 3x. Ideal: > 5x"),
    ("Payback Period (months)",   ["=C4/(C5*C6)", "=D4/(D5*D6)", "=E4/(E5*E6)", "=F4/(F5*F6)"],
                                                                "0.0",      "Months to recover CAC"),
    ("12-Month Gross Profit ($)", ["=C5*12*C5*C6*(1-(1-C7)^12)/C7",
                                   "=D5*12*D5*D6*(1-(1-D7)^12)/D7",
                                   "=E5*12*E5*E6*(1-(1-E7)^12)/E7",
                                   "=F5*12*F5*F6*(1-(1-F7)^12)/F7"],
                                                                "#,##0",    "Gross profit per cohort over 12 months"),
]

for r, (name, values, fmt, note) in enumerate(PARAMS, start=4):
    label(ws1, r, 1, name, bold=(r in (4, 11)))
    bg = LGRAY if r % 2 == 0 else WHITE
    for c, v in enumerate(values, start=2):
        cell = val_cell(ws1, r, c, v, fmt, bg)
        cell.border = thin_border()

# Notes column
ws1.column_dimensions["G"].width = 45
ws1.cell(row=3, column=7, value="Notes").font = Font(bold=True, color=BLACK)
for r, (_, _, _, note) in enumerate(PARAMS, start=4):
    c = ws1.cell(row=r, column=7, value=note)
    c.font = Font(color="FF666666", size=9, italic=True)

# Freeze header
ws1.freeze_panes = "B4"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: UNIT ECONOMICS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Unit Economics")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F"]: ws2.column_dimensions[col].width = 15
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A1:F1")
hdr(ws2, 1, 1, "Unit Economics Summary — All Channels", NAVY, WHITE, size=13)

# Headers
hdr(ws2, 3, 1, "Metric",           LGRAY, BLACK, bold=True)
for i, (ch, clr) in enumerate(zip(CHANNELS, CH_COLORS)):
    hdr(ws2, 3, i+2, ch, clr, WHITE)
hdr(ws2, 3, 6, "Total / Avg",      NAVY, WHITE)

METRICS = [
    ("Monthly Budget ($)",         "=Inputs!B4",  "=Inputs!C4",  "=Inputs!D4",  "=Inputs!E4",
     "=SUM(B4:E4)", "#,##0"),
    ("New Customers / Month",      "=Inputs!B5",  "=Inputs!C5",  "=Inputs!D5",  "=Inputs!E5",
     "=SUM(B5:E5)", "#,##0"),
    ("CAC ($)",                    "=Inputs!B6",  "=Inputs!C6",  "=Inputs!D6",  "=Inputs!E6",
     "=SUMPRODUCT(B4:E4,B6:E6)/SUM(B4:E4)", "#,##0.00"),
    ("LTV ($)",                    "=Inputs!B8",  "=Inputs!C8",  "=Inputs!D8",  "=Inputs!E8",
     "=SUMPRODUCT(B5:E5,B8:E8)/SUM(B5:E5)", "#,##0"),
    ("LTV / CAC",                  "=Inputs!B9",  "=Inputs!C9",  "=Inputs!D9",  "=Inputs!E9",
     "=F8/F6", "0.0x"),
    ("Payback Period (months)",    "=Inputs!B10", "=Inputs!C10", "=Inputs!D10", "=Inputs!E10",
     "=SUMPRODUCT(B5:E5,B10:E10)/SUM(B5:E5)", "0.0"),
    ("Monthly Gross Profit ($)",   "=B5*Inputs!B5*Inputs!B6",
                                   "=C5*Inputs!C5*Inputs!C6",
                                   "=D5*Inputs!D5*Inputs!D6",
                                   "=E5*Inputs!E5*Inputs!E6",
     "=SUM(B11:E11)", "#,##0"),
    ("Budget Share (%)",           "=B4/F4",      "=C4/F4",      "=D4/F4",      "=E4/F4",
     "=SUM(B12:E12)", "0.0%"),
    ("Revenue Share (%)",          "=B5*Inputs!B5/SUM(B5:E5*Inputs!B5:Inputs!E5)",
                                   "=C5*Inputs!C5/SUM(B5:E5*Inputs!B5:Inputs!E5)",
                                   "=D5*Inputs!D5/SUM(B5:E5*Inputs!B5:Inputs!E5)",
                                   "=E5*Inputs!E5/SUM(B5:E5*Inputs!B5:Inputs!E5)",
     "=SUM(B13:E13)", "0.0%"),
]

for r, row_data in enumerate(METRICS, start=4):
    name = row_data[0]
    vals = row_data[1:5]
    total = row_data[5]
    fmt  = row_data[6]
    bg   = LGRAY if r % 2 == 0 else WHITE

    label(ws2, r, 1, name, bold=True)
    for c, v in enumerate(vals, start=2):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right")
    tot = ws2.cell(row=r, column=6, value=total)
    tot.number_format = fmt
    tot.font = Font(bold=True)
    tot.fill = PatternFill("solid", fgColor=LGRAY)
    tot.border = thin_border()
    tot.alignment = Alignment(horizontal="right")

ws2.freeze_panes = "B4"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 28
for col in ["B","C","D"]: ws3.column_dimensions[col].width = 20
ws3.row_dimensions[1].height = 36

ws3.merge_cells("A1:D1")
hdr(ws3, 1, 1, "Budget Reallocation Scenarios — Total $16 500 / Month", NAVY, WHITE, size=13)

hdr(ws3, 3, 1, "Channel / Metric", LGRAY, BLACK, bold=True)
hdr(ws3, 3, 2, "Current",          ORANGE, WHITE)
hdr(ws3, 3, 3, "Optimised",        TEAL, WHITE)
hdr(ws3, 3, 4, "Aggressive SEO",   NAVY, WHITE)

SCENARIOS = [
    # (label, current, optimised, aggressive_seo)
    ("BUDGET ALLOCATION", None, None, None),
    ("Google Ads ($)",     8_000,  7_000,  6_000),
    ("Facebook Ads ($)",   6_000,  3_500,  2_000),
    ("SEO ($)",            2_000,  5_000,  7_500),
    ("Email ($)",            500,  1_000,  1_000),
    ("Total ($)",         16_500, 16_500, 16_500),
    ("", None, None, None),
    ("PROJECTED OUTCOMES", None, None, None),
    ("New Customers / Mo", 436,    490,    560),
    ("Blended CAC ($)",    37.8,   33.7,   29.5),
    ("Blended LTV ($)",    616,    720,    820),
    ("LTV/CAC Ratio",      16.3,   21.4,   27.8),
    ("Monthly Gross Profit ($)", 9_700, 11_200, 13_100),
    ("12-Month incremental ($)", None, 18_000, 40_800),
]

FMT_MAP = {
    "Google Ads ($)": "#,##0", "Facebook Ads ($)": "#,##0",
    "SEO ($)": "#,##0", "Email ($)": "#,##0", "Total ($)": "#,##0",
    "New Customers / Mo": "#,##0", "Blended CAC ($)": "#,##0.00",
    "Blended LTV ($)": "#,##0", "LTV/CAC Ratio": "0.0x",
    "Monthly Gross Profit ($)": "#,##0", "12-Month incremental ($)": "#,##0",
}

for r, (name, cur, opt, agg) in enumerate(SCENARIOS, start=4):
    if name in ("BUDGET ALLOCATION", "PROJECTED OUTCOMES", ""):
        if name:
            c = ws3.cell(row=r, column=1, value=name)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            ws3.merge_cells(f"A{r}:D{r}")
        continue

    label(ws3, r, 1, name)
    bg = LGRAY if r % 2 == 0 else WHITE
    fmt = FMT_MAP.get(name, "General")
    for c, v in zip([2, 3, 4], [cur, opt, agg]):
        if v is not None:
            cell = ws3.cell(row=r, column=c, value=v)
            cell.number_format = fmt
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right")
            if name == "Total ($)":
                cell.font = Font(bold=True)
            if name == "12-Month incremental ($)" and v:
                cell.font = Font(bold=True, color=TEAL if c > 2 else BLACK)

ws3.freeze_panes = "B4"

wb.save(OUT)
print(f"✓ Saved {OUT}")
